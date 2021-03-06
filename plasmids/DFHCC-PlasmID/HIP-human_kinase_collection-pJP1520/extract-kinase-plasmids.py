import os
import gzip
import openpyxl
import Bio, Bio.Seq, Bio.Alphabet
from lxml import etree
import pandas as pd
import imp
project_tools = imp.load_source('project_tools', '../../../project_pylib/project_tools.py')
sequnwrap = project_tools.sequnwrap

plasmid_library_dir = '.'
Harvard_plasmid_library_filepath = os.path.join(plasmid_library_dir, 'Mehle_Kinase_VS1_pJP1520_new_plates.xlsx')

# ========
# Command-line args
# ========

database_path = os.path.join('..', '..', '..', 'kinome-database', 'database.xml.gz')

desired_taxid = 9606
ignore_clones = ['HsCD00038286', 'HsCD00037967'] # both of these have deletions in the kinase catalytic domain

# ==========
# Parse plasmid library spreadsheet
# ==========

wb = openpyxl.load_workbook(Harvard_plasmid_library_filepath)
sheet_ranges = wb.get_sheet_by_name(name = 'Kinase_VS1_pJP1520_new_plates')
nrows = sheet_ranges.get_highest_row()
plasmid_df = {
'cloneID':[],
'dna_seq':[],
'dna_orf_seq':[],
'aa_seq':[],
'NCBI_GeneID':[],
'Symbol':[],
'plateID':[],
'well_pos':[],
}

for row in range(2,nrows):
    cloneID = sheet_ranges.cell('E%d' % row).value    # type(int)
    if cloneID in ignore_clones:
        continue
    NCBI_GeneID = sheet_ranges.cell('H%d' % row).value    # type(int)
    Symbol = sheet_ranges.cell('I%d' % row).value
    dna_seq = sheet_ranges.cell('L%d' % row).value.upper()
    if len(dna_seq) % 3 != 0:
        print 'WARNING: length of DNA sequence not divisible by 3, for plasmid with Gene Symbol %s and NCBI Gene ID %s' % (Symbol, NCBI_GeneID)
    # Translate to DNA sequence (don't include stop codon)
    aa_seq = Bio.Seq.Seq(dna_seq, Bio.Alphabet.generic_dna).translate(to_stop=True)

    # get DNA ORF (assume starts at index 0)
    dna_orf_seq = dna_seq
    for n in range(0, len(dna_seq), 3):
        if dna_seq[n:n+3] in ['TAG', 'TAA', 'TGA']:
           dna_orf_seq = dna_seq[:n]
           break

    plateID = sheet_ranges.cell('A%d' % row).value
    well_pos = sheet_ranges.cell('C%d' % row).value

    plasmid_df['cloneID'].append(cloneID)
    plasmid_df['dna_seq'].append(dna_seq)
    plasmid_df['dna_orf_seq'].append(dna_orf_seq)
    plasmid_df['aa_seq'].append(aa_seq)
    plasmid_df['NCBI_GeneID'].append(NCBI_GeneID)
    plasmid_df['Symbol'].append(Symbol)
    plasmid_df['plateID'].append(plateID)
    plasmid_df['well_pos'].append(well_pos)

#plasmid_NCBI_GeneIDs = plasmid_aa_seqs.keys()

plasmid_df = pd.DataFrame(plasmid_df)

with gzip.open(database_path) as database_file:
    DB_root = etree.parse(database_file).getroot()


# To be used to construct a pandas DataFrame
data_fields = ['cloneID', 'NCBI_GeneID', 'orig_gene_symbol', 'UniProtAC', 'UniProt_entry_name', 'UniProt_family', 'construct_dna_seq', 'construct_dna_orf_seq', 'construct_aa_seq', 'plateID', 'well_pos']
output_data = pd.DataFrame( [['None'] * len(data_fields)] * len(plasmid_df), columns=data_fields)

#DB_gene_name_nodes = [ gene_name_node for gene_name_node in DB_root.findall('entry/UniProt/gene_names/gene_name') ]

# ===========
# Iterate through plasmids
# ===========

for p in plasmid_df.index:
    cloneID = plasmid_df['cloneID'][p]
    construct_dna_seq = plasmid_df['dna_seq'][p]
    construct_dna_orf_seq = plasmid_df['dna_orf_seq'][p]
    construct_aa_seq = plasmid_df['aa_seq'][p]
    plasmid_NCBI_GeneID = plasmid_df['NCBI_GeneID'][p]
    plasmid_Symbol = plasmid_df['Symbol'][p]
    plateID = plasmid_df['plateID'][p]
    well_pos = plasmid_df['well_pos'][p]

    output_data['cloneID'][p] = cloneID
    output_data['NCBI_GeneID'][p] = plasmid_NCBI_GeneID
    output_data['orig_gene_symbol'][p] = plasmid_Symbol
    output_data['construct_dna_seq'][p] = construct_dna_seq
    output_data['construct_dna_orf_seq'][p] = construct_dna_orf_seq
    output_data['construct_aa_seq'][p] = construct_aa_seq
    output_data['plateID'][p] = plateID
    output_data['well_pos'][p] = well_pos


    # find matching DB entry via NCBI GeneID
    DB_entry = DB_root.find('entry/UniProt[@NCBI_taxID="%d"]/../NCBI_Gene/entry[@ID="%s"]/../..' % (desired_taxid, plasmid_NCBI_GeneID))
    if DB_entry == None:
        print 'Matching DB entry not found for Gene ID %s cloneID %s Symbol %s' % (plasmid_NCBI_GeneID, cloneID, plasmid_Symbol)

        # check if the plasmid gene symbol matches in any in the DB
        matching_DB_gene_names = DB_root.xpath('entry/UniProt/gene_names/gene_name[text()="%s"]' % plasmid_Symbol)
        if len(matching_DB_gene_names) > 0:
            for matching_DB_gene_name in matching_DB_gene_names:
                matching_DB_UniProt_entry_name = matching_DB_gene_name.getparent().getparent().get('entry_name')
                print 'NOTE: plasmid gene symbol %s for clone %s appears to match a gene symbol in the DB entry %s' % (plasmid_Symbol, cloneID, matching_DB_UniProt_entry_name)

        continue


    DB_UniProt_node = DB_entry.find('UniProt')
    DB_domains = DB_UniProt_node.findall('domains/domain[@targetID]')
    UniProtAC = DB_UniProt_node.get('AC')
    UniProt_entry_name = DB_UniProt_node.get('entry_name')
    UniProt_family = DB_UniProt_node.get('family')
    UniProt_canonseq = sequnwrap( DB_UniProt_node.find('isoforms/canonical_isoform/sequence').text )

    output_data['UniProtAC'][p] = UniProtAC
    output_data['UniProt_entry_name'][p] = UniProt_entry_name
    output_data['UniProt_family'][p] = UniProt_family

    #break

# construct pandas DataFrame and write to csv

output_data.set_index('cloneID', inplace=True)
output_data.to_csv('plasmid-data.csv')

